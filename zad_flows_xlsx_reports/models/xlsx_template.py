# -*- coding: utf-8 -*-
from odoo import api, fields, models, _
from odoo.exceptions import UserError, ValidationError
import base64
from io import BytesIO
from datetime import datetime, date
from jinja2 import Environment, StrictUndefined
from num2words import num2words
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage
import re

def spelled_out(value, lang='en'):
    try:
        return num2words(value, lang=lang)
    except Exception:
        return str(value)

def formatdate(value, fmt='%Y-%m-%d'):
    if not value:
        return ''
    if isinstance(value, (datetime, date)):
        return value.strftime(fmt)
    try:
        # try parse from string
        dt = datetime.fromisoformat(str(value))
        return dt.strftime(fmt)
    except Exception:
        return str(value)

def convert_currency(amount, currency, precision=None):
    if currency and hasattr(currency, 'round'):
        return currency.round(amount)
    if precision is None:
        precision = 2
    return round(amount or 0.0, precision)

class ZadXlsxTemplate(models.Model):
    _name = 'zad.xlsx.template'
    _description = 'Zad Flows XLSX Template'

    name = fields.Char(required=True)
    model_id = fields.Many2one('ir.model', string='Applies to Model', required=True)
    template_file = fields.Binary(string='XLSX Template (.xlsx)', required=True, attachment=True)
    template_filename = fields.Char()
    active = fields.Boolean(default=True)
    note = fields.Text()

    def _render_context(self, record):
        # Build a safe context for Jinja
        ctx = {
            'docs': record,
            'spelled_out': spelled_out,
            'formatdate': formatdate,
            'convert_currency': lambda amount, currency=None, precision=None: convert_currency(amount, currency, precision),
        }
        return ctx

    def _render_cell(self, value, context):
        if not isinstance(value, str):
            return value
        # Render Jinja expressions in the cell
        env = Environment(autoescape=False, undefined=StrictUndefined)
        template = env.from_string(value)
        try:
            return template.render(context)
        except Exception as e:
            # keep raw if rendering failed
            return value

    def _apply_images(self, ws, context):
        # Replace tags like {% img docs.image_field %} with the actual image, if the field is binary (base64)
        # We scan cell values; if found, we clear the cell and place an image anchored to the cell.
        pattern = re.compile(r"{%\s*img\s+([^%]+)\s*%}")
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    m = pattern.search(cell.value)
                    if m:
                        expr = m.group(1).strip()
                        # Evaluate expression against context by Jinja
                        from jinja2 import Environment, StrictUndefined
                        env = Environment(autoescape=False, undefined=StrictUndefined)
                        try:
                            val = env.from_string("{{ " + expr + " }}").render(context)
                        except Exception:
                            val = None
                        if val:
                            try:
                                raw = None
                                if isinstance(val, (bytes, bytearray)):
                                    raw = val
                                elif isinstance(val, str):
                                    # try base64
                                    try:
                                        raw = base64.b64decode(val)
                                    except Exception:
                                        raw = None
                                if raw:
                                    bio = BytesIO(raw)
                                    img = PILImage.open(bio)
                                    tmp = BytesIO()
                                    img.save(tmp, format='PNG')
                                    tmp.seek(0)
                                    xlimg = XLImage(tmp)
                                    ws.add_image(xlimg, cell.coordinate)
                                    cell.value = None
                            except Exception:
                                pass

    def generate_xlsx_bytes(self, record):
        self.ensure_one()
        if not self.template_file:
            raise UserError(_('No template file set.'))
        data = base64.b64decode(self.template_file)
        wb = load_workbook(filename=BytesIO(data))
        context = self._render_context(record)
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    cell.value = self._render_cell(cell.value, context)
            # images last
            self._apply_images(ws, context)
        out = BytesIO()
        wb.save(out)
        return out.getvalue()

class IrActionsReport(models.Model):
    _inherit = 'ir.actions.report'

    zad_template_id = fields.Many2one('zad.xlsx.template', string='Zad XLSX Template')

    def _render_xlsx(self, report_ref, res_ids):
        self.ensure_one()
        if not self.zad_template_id:
            raise UserError(_('No Zad XLSX Template linked to this report.'))
        if len(res_ids) != 1:
            raise UserError(_('This report expects exactly one record.'))
        record = self.env[self.zad_template_id.model_id.model].browse(res_ids[0])
        content = self.zad_template_id.generate_xlsx_bytes(record)
        return content, 'xlsx'
